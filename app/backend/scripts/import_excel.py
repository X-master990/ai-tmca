"""把 110-115年 總表.xlsx 匯入 records 表

用法:
    # dry-run（預設，不寫 DB）
    docker compose exec backend python scripts/import_excel.py /var/tmca/import/總表.xlsx

    # 真寫入
    docker compose exec backend python scripts/import_excel.py /var/tmca/import/總表.xlsx --commit
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models import Category, Record
from app.utils.roc_date import roc_to_ad


# ────────────────────────────────────────────────────────────────
# Column mapping DSL
# ────────────────────────────────────────────────────────────────
# 每個 column 一個 tuple: (col_idx, kind, key, dtype)
#   kind:
#     "col"   → Record.<key>
#     "extra" → extra[<key>]
#     "ps_y/ps_m/ps_d" → period_start 的年/月/日（民國年）
#     "pe_y/pe_m/pe_d" → period_end 的年/月/日（民國年）
#     "skip"  → 不處理
#   dtype: "date" | "int" | "text" | "bool" | None
INVOICE_AMOUNT_COMMON = [
    (1, "skip", None, None),
    (2, "col", "issued_date", "date"),
    (3, "col", "note", "text"),
    (4, "col", "cert_no", "text"),
    (5, "col", "invoice_date", "date"),
    (6, "col", "invoice_type", "text"),
    (7, "col", "invoice_title", "text"),
    (8, "col", "tax_id", "text"),
    (9, "col", "invoice_no", "text"),
    (10, "col", "amount", "int"),
]

KARAOKE_TAIL = [
    (11, "col", "source", "text"),
    (12, "col", "officer", "text"),
    (13, "col", "action_type", "text"),
    (14, "col", "apply_date", "date"),
    (15, "col", "applicant_name", "text"),
    (16, "col", "applicant_id", "text"),
    (17, "col", "applicant_mobile", "text"),
    (18, "col", "applicant_phone", "text"),
    (19, "col", "applicant_fax", "text"),
    (20, "col", "holder_name", "text"),
    (21, "col", "holder_type", "text"),
    (22, "col", "use_zip", "text"),
    (23, "col", "use_address", "text"),
    (24, "col", "onsite_name", "text"),
    (25, "col", "onsite_mobile", "text"),
    (26, "col", "onsite_phone", "text"),
    (27, "col", "onsite_ext", "text"),
    (28, "col", "onsite_fax", "text"),
    (29, "col", "qty", "int"),
    (30, "col", "brand", "text"),
    (31, "col", "serial_no", "text"),
    (32, "ps_y", None, "int"),
    (33, "ps_m", None, "int"),
    (34, "ps_d", None, "int"),
    (35, "pe_y", None, "int"),
    (36, "pe_m", None, "int"),
    (37, "pe_d", None, "int"),
    (38, "col", "mail_zip", "text"),
    (39, "col", "mail_address", "text"),
    (40, "col", "mail_recipient", "text"),
    (41, "col", "mail_phone", "text"),
]

# 每個 sheet 的設定
SHEET_CONFIG: dict[str, dict[str, Any]] = {
    "電腦伴唱機": {
        "category_code": "COMPUTER_KARAOKE",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + KARAOKE_TAIL,
    },
    "社區管委會": {
        "category_code": "COMMUNITY_BOARD",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + KARAOKE_TAIL,
    },
    "公益伴唱機": {
        "category_code": "PUBLIC_KARAOKE",
        "data_start_row": 3,
        # 42 欄，最後一欄空白
        "columns": INVOICE_AMOUNT_COMMON + KARAOKE_TAIL + [(42, "skip", None, None)],
    },
    "自助KTV": {
        "category_code": "SELF_SERVICE_KTV",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "source", "text"),
            (12, "col", "officer", "text"),
            (13, "col", "action_type", "text"),
            (14, "col", "apply_date", "date"),
            (15, "col", "applicant_name", "text"),
            (16, "col", "applicant_id", "text"),
            (17, "col", "applicant_mobile", "text"),
            (18, "col", "applicant_phone", "text"),
            (19, "col", "applicant_fax", "text"),
            (20, "col", "holder_name", "text"),
            (21, "col", "onsite_name", "text"),
            (22, "col", "onsite_mobile", "text"),
            (23, "col", "onsite_phone", "text"),
            (24, "col", "onsite_ext", "text"),
            (25, "col", "onsite_fax", "text"),
            (26, "col", "use_zip", "text"),
            (27, "col", "use_address", "text"),
            (28, "col", "qty", "int"),
            (29, "ps_y", None, "int"),
            (30, "ps_m", None, "int"),
            (31, "ps_d", None, "int"),
            (32, "pe_y", None, "int"),
            (33, "pe_m", None, "int"),
            (34, "pe_d", None, "int"),
            (35, "col", "mail_zip", "text"),
            (36, "col", "mail_address", "text"),
            (37, "col", "mail_recipient", "text"),
            (38, "col", "mail_phone", "text"),
        ],
    },
    "街頭藝人": {
        "category_code": "STREET_ARTIST",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "source", "text"),
            (12, "col", "officer", "text"),
            (13, "col", "action_type", "text"),
            (14, "col", "apply_date", "date"),
            (15, "col", "applicant_name", "text"),
            (16, "col", "applicant_mobile", "text"),
            (17, "col", "applicant_phone", "text"),
            (18, "col", "applicant_fax", "text"),
            (19, "extra", "email", "text"),
            (20, "col", "holder_name", "text"),
            (21, "extra", "cert_issuer", "text"),
            (22, "extra", "street_cert_no", "text"),
            (23, "extra", "street_cert_expiry", "text"),  # 民國年 string 保留原樣
            (24, "col", "use_address", "text"),
            (25, "ps_y", None, "int"),
            (26, "ps_m", None, "int"),
            (27, "ps_d", None, "int"),
            (28, "pe_y", None, "int"),
            (29, "pe_m", None, "int"),
            (30, "pe_d", None, "int"),
            (31, "col", "mail_zip", "text"),
            (32, "col", "mail_address", "text"),
            (33, "col", "mail_recipient", "text"),
            (34, "col", "mail_phone", "text"),
        ],
    },
    "交通運輸工具": {
        "category_code": "TRANSPORT",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "source", "text"),
            (12, "col", "officer", "text"),
            (13, "col", "action_type", "text"),
            (14, "col", "apply_date", "date"),
            (15, "col", "applicant_name", "text"),
            (16, "col", "applicant_mobile", "text"),
            (17, "col", "applicant_phone", "text"),
            (18, "col", "applicant_fax", "text"),
            (19, "col", "holder_name", "text"),
            (20, "col", "use_zip", "text"),
            (21, "col", "use_address", "text"),
            (22, "col", "onsite_name", "text"),
            (23, "col", "onsite_mobile", "text"),
            (24, "col", "onsite_phone", "text"),
            (25, "col", "onsite_ext", "text"),
            (26, "col", "onsite_fax", "text"),
            (27, "col", "qty", "int"),
            (28, "col", "serial_no", "text"),
            (29, "ps_y", None, "int"),
            (30, "ps_m", None, "int"),
            (31, "ps_d", None, "int"),
            (32, "pe_y", None, "int"),
            (33, "pe_m", None, "int"),
            (34, "pe_d", None, "int"),
            (35, "col", "mail_zip", "text"),
            (36, "col", "mail_address", "text"),
            (37, "col", "mail_recipient", "text"),
            (38, "col", "mail_phone", "text"),
        ],
    },
    "單場次表演": {
        "category_code": "SINGLE_EVENT",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "apply_date", "date"),
            (12, "col", "holder_name", "text"),
            (13, "extra", "holder_tax_id", "text"),
            (14, "col", "applicant_phone", "text"),
            (15, "col", "applicant_fax", "text"),
            (16, "col", "use_address", "text"),
            (17, "extra", "event_name", "text"),
            (18, "extra", "songs", "text"),
            (19, "extra", "song_count", "text"),
            (20, "extra", "venue", "text"),
            (21, "extra", "venue_address", "text"),
            (22, "col", "qty", "int"),
            (23, "extra", "audience_size", "int"),
            (24, "ps_y", None, "int"),
            (25, "ps_m", None, "int"),
            (26, "ps_d", None, "int"),
            (27, "pe_y", None, "int"),
            (28, "pe_m", None, "int"),
            (29, "pe_d", None, "int"),
            (30, "extra", "contact_org", "text"),
            (31, "col", "onsite_name", "text"),
            (32, "extra", "contact_title", "text"),
            (33, "col", "onsite_mobile", "text"),
            (34, "col", "onsite_phone", "text"),
            (35, "col", "onsite_ext", "text"),
            (36, "col", "onsite_fax", "text"),
            (37, "extra", "contact_email", "text"),
            (38, "col", "mail_zip", "text"),
            (39, "col", "mail_address", "text"),
            (40, "col", "mail_recipient", "text"),
            (41, "col", "mail_phone", "text"),
        ],
    },
    "公開傳輸": {
        "category_code": "PUBLIC_TRANSMIT",
        "data_start_row": 2,  # 不一樣！
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "apply_date", "date"),
            (12, "col", "holder_name", "text"),
            (13, "col", "applicant_name", "text"),
            (14, "col", "onsite_name", "text"),
            (15, "col", "applicant_phone", "text"),
            (16, "col", "applicant_fax", "text"),
            (17, "extra", "email", "text"),
            (18, "col", "use_zip", "text"),
            (19, "col", "use_address", "text"),
            (20, "col", "holder_type", "text"),
            (21, "extra", "has_revenue", "text"),
            (22, "extra", "platform_name", "text"),
            (23, "extra", "platform_url", "text"),
            (24, "extra", "songs", "text"),
            (25, "skip", None, None),
            (26, "ps_y", None, "int"),
            (27, "ps_m", None, "int"),
            (28, "ps_d", None, "int"),
            (29, "pe_y", None, "int"),
            (30, "pe_m", None, "int"),
            (31, "pe_d", None, "int"),
            (32, "col", "mail_zip", "text"),
            (33, "col", "mail_address", "text"),
            (34, "col", "mail_recipient", "text"),
            (35, "skip", None, None),
        ],
    },
    "告別式": {
        "category_code": "FUNERAL",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "apply_date", "date"),
            (12, "col", "applicant_name", "text"),
            (13, "col", "applicant_mobile", "text"),
            (14, "col", "applicant_phone", "text"),
            (15, "col", "applicant_fax", "text"),
            (16, "col", "holder_name", "text"),
            (17, "extra", "ceremony_name", "text"),
            (18, "extra", "songs", "text"),
            (19, "extra", "language", "text"),
            (20, "extra", "song_count", "text"),
            (21, "extra", "venue", "text"),
            (22, "col", "use_address", "text"),
            (23, "col", "qty", "int"),
            (24, "ps_y", None, "int"),
            (25, "ps_m", None, "int"),
            (26, "ps_d", None, "int"),
            (27, "pe_y", None, "int"),
            (28, "pe_m", None, "int"),
            (29, "pe_d", None, "int"),
            (30, "extra", "funeral_company", "text"),
            (31, "col", "onsite_name", "text"),
            (32, "col", "onsite_mobile", "text"),
            (33, "col", "onsite_phone", "text"),
            (34, "col", "onsite_fax", "text"),
            (35, "extra", "contact_email", "text"),
            (36, "col", "mail_zip", "text"),
            (37, "col", "mail_address", "text"),
            (38, "col", "mail_recipient", "text"),
            (39, "col", "mail_phone", "text"),
        ],
    },
    "坪數-顯示器": {
        "category_code": "AREA_DISPLAY",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "apply_date", "date"),
            (12, "col", "holder_name", "text"),
            (13, "col", "applicant_phone", "text"),
            (14, "extra", "applicant_ext", "text"),
            (15, "col", "applicant_fax", "text"),
            (16, "col", "applicant_name", "text"),
            (17, "extra", "email", "text"),
            (18, "extra", "event_name", "text"),
            (19, "col", "use_address", "text"),
            (20, "extra", "floor_area", "int"),
            (21, "col", "qty", "int"),
            (22, "ps_y", None, "int"),
            (23, "ps_m", None, "int"),
            (24, "ps_d", None, "int"),
            (25, "pe_y", None, "int"),
            (26, "pe_m", None, "int"),
            (27, "pe_d", None, "int"),
            (28, "extra", "songs", "text"),
            (29, "extra", "language", "text"),
            (30, "col", "mail_zip", "text"),
            (31, "col", "mail_address", "text"),
            (32, "col", "mail_recipient", "text"),
            (33, "col", "mail_phone", "text"),
        ],
    },
    "大廳-宴會廳-客房": {
        "category_code": "HALL_ROOM",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "action_type", "text"),
            (12, "col", "apply_date", "date"),
            (13, "col", "holder_name", "text"),
            (14, "col", "applicant_name", "text"),
            (15, "col", "applicant_phone", "text"),
            (16, "extra", "applicant_ext", "text"),
            (17, "col", "applicant_fax", "text"),
            (18, "col", "onsite_name", "text"),
            (19, "col", "onsite_mobile", "text"),
            (20, "extra", "email", "text"),
            (21, "col", "use_address", "text"),
            (22, "col", "holder_type", "text"),
            (23, "extra", "floor_area", "int"),
            (24, "col", "qty", "int"),
            (25, "ps_y", None, "int"),
            (26, "ps_m", None, "int"),
            (27, "ps_d", None, "int"),
            (28, "pe_y", None, "int"),
            (29, "pe_m", None, "int"),
            (30, "pe_d", None, "int"),
            (31, "col", "mail_zip", "text"),
            (32, "col", "mail_address", "text"),
            (33, "col", "mail_recipient", "text"),
            (34, "col", "mail_phone", "text"),
        ],
    },
    "競選活動": {
        "category_code": "ELECTION",
        "data_start_row": 3,
        "columns": INVOICE_AMOUNT_COMMON + [
            (11, "col", "apply_date", "date"),
            (12, "col", "holder_name", "text"),
            (13, "extra", "holder_tax_id", "text"),
            (14, "col", "applicant_phone", "text"),
            (15, "col", "applicant_fax", "text"),
            (16, "col", "use_address", "text"),
            (17, "extra", "event_name", "text"),
            (18, "extra", "venue", "text"),
            (19, "extra", "venue_address", "text"),
            (20, "col", "holder_type", "text"),
            (21, "col", "qty", "int"),
            (22, "extra", "songs", "text"),
            (23, "extra", "language", "text"),
            (24, "extra", "song_count", "text"),
            (25, "ps_y", None, "int"),
            (26, "ps_m", None, "int"),
            (27, "ps_d", None, "int"),
            (28, "pe_y", None, "int"),
            (29, "pe_m", None, "int"),
            (30, "pe_d", None, "int"),
            (31, "extra", "contact_org", "text"),
            (32, "col", "onsite_name", "text"),
            (33, "extra", "contact_title", "text"),
            (34, "col", "onsite_mobile", "text"),
            (35, "col", "onsite_phone", "text"),
            (36, "col", "onsite_ext", "text"),
            (37, "col", "onsite_fax", "text"),
            (38, "extra", "contact_email", "text"),
            (39, "col", "mail_zip", "text"),
            (40, "col", "mail_address", "text"),
            (41, "col", "mail_recipient", "text"),
            (42, "col", "mail_phone", "text"),
        ],
    },
}


# ────────────────────────────────────────────────────────────────
# 值清洗
# ────────────────────────────────────────────────────────────────
def _clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def _to_int(v: Any) -> int | None:
    v = _clean(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _to_date(v: Any) -> date | None:
    """民國年字串 → date。已是 datetime/date 直接回。"""
    v = _clean(v)
    if v is None:
        return None
    if isinstance(v, date):
        return v
    if hasattr(v, "date"):
        return v.date()
    return roc_to_ad(str(v))


def _to_text(v: Any) -> str | None:
    v = _clean(v)
    if v is None:
        return None
    return str(v)


def _to_period_date(y: Any, m: Any, d: Any) -> date | None:
    """民國年 y/m/d 三欄 → date。任一缺值或不合法 → None。"""
    yi, mi, di = _to_int(y), _to_int(m), _to_int(d)
    if yi is None or mi is None or di is None:
        return None
    try:
        return date(yi + 1911, mi, di)
    except ValueError:
        return None


# ────────────────────────────────────────────────────────────────
# 主流程
# ────────────────────────────────────────────────────────────────
def parse_row(
    row: tuple,
    columns_cfg: list[tuple],
) -> tuple[dict, dict, dict]:
    """回傳 (record_kwargs, extra_dict, period_parts)"""
    rec: dict[str, Any] = {}
    extra: dict[str, Any] = {}
    parts: dict[str, Any] = {}

    for col_idx, kind, key, dtype in columns_cfg:
        v = row[col_idx - 1] if col_idx - 1 < len(row) else None
        if kind == "skip":
            continue
        if kind in ("ps_y", "ps_m", "ps_d", "pe_y", "pe_m", "pe_d"):
            parts[kind] = v
            continue

        if dtype == "date":
            v = _to_date(v)
        elif dtype == "int":
            v = _to_int(v)
        else:
            v = _to_text(v)

        if v is None:
            continue

        if kind == "col":
            rec[key] = v
        elif kind == "extra":
            extra[key] = v

    return rec, extra, parts


def is_empty_row(row: tuple) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def process_sheet(ws, cfg: dict, db: Session | None, dry_run: bool) -> dict:
    code = cfg["category_code"]
    start = cfg["data_start_row"]
    cols_cfg = cfg["columns"]

    stats = {
        "sheet": ws.title,
        "category": code,
        "total_rows": 0,
        "success": 0,
        "skipped_empty": 0,
        "skipped_thin": 0,
        "errors": [],
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=start, values_only=True), start=start):
        stats["total_rows"] += 1
        if is_empty_row(row):
            stats["skipped_empty"] += 1
            continue

        try:
            rec_kw, extra, parts = parse_row(row, cols_cfg)
            # 「有意義」門檻：至少要有 cert_no / holder_name / invoice_no / amount 之一
            if not any(rec_kw.get(k) for k in ("cert_no", "holder_name", "invoice_no", "amount")):
                stats["skipped_thin"] += 1
                continue

            period_start = _to_period_date(parts.get("ps_y"), parts.get("ps_m"), parts.get("ps_d"))
            period_end = _to_period_date(parts.get("pe_y"), parts.get("pe_m"), parts.get("pe_d"))
            if period_start:
                rec_kw["period_start"] = period_start
            if period_end:
                rec_kw["period_end"] = period_end

            rec_kw["category_code"] = code
            if extra:
                rec_kw["extra"] = extra

            # 自動計算 issuance_status: 有發票號碼 → 綠
            if rec_kw.get("invoice_no"):
                rec_kw["issuance_status"] = "綠"
            else:
                rec_kw["issuance_status"] = "紅"

            if not dry_run and db is not None:
                rec = Record(**rec_kw)
                db.add(rec)

            stats["success"] += 1
        except Exception as e:
            stats["errors"].append({"row": row_idx, "cert_no": rec_kw.get("cert_no") if 'rec_kw' in dir() else None, "error": str(e)})

    if not dry_run and db is not None:
        db.commit()

    return stats


def main():
    parser = argparse.ArgumentParser(description="匯入 110-115 年總表 Excel → records")
    parser.add_argument("xlsx_path", type=str, help="xlsx 路徑")
    parser.add_argument("--commit", action="store_true", help="實際寫入 DB（預設為 dry-run）")
    parser.add_argument("--only", type=str, default=None, help="只跑指定 sheet（中文名）")
    parser.add_argument("--purge", action="store_true", help="開始前先清空 records（小心！）")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"❌ 找不到檔案: {xlsx_path}")
        sys.exit(1)

    print(f"📂 來源: {xlsx_path}")
    print(f"🎯 模式: {'COMMIT (寫入 DB)' if args.commit else 'DRY-RUN (不寫入)'}")
    if args.only:
        print(f"🔍 只處理 sheet: {args.only}")
    print()

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    db: Session | None = None
    if args.commit:
        db = SessionLocal()
        # 確認 categories 存在
        existing = {c.code for c in db.query(Category).all()}
        missing = [cfg["category_code"] for cfg in SHEET_CONFIG.values() if cfg["category_code"] not in existing]
        if missing:
            print(f"❌ categories 缺資料: {missing}")
            print("   先跑: python scripts/seed_categories.py")
            sys.exit(1)
        if args.purge:
            n = db.query(Record).delete()
            db.commit()
            print(f"🗑️  已清空 records ({n} 筆)\n")

    all_stats: list[dict] = []
    try:
        for sheet_name in wb.sheetnames:
            if args.only and sheet_name != args.only:
                continue
            if sheet_name not in SHEET_CONFIG:
                print(f"⚠️  跳過未定義 sheet: {sheet_name}")
                continue
            ws = wb[sheet_name]
            stats = process_sheet(ws, SHEET_CONFIG[sheet_name], db, dry_run=not args.commit)
            all_stats.append(stats)
            print(f"  {sheet_name:>16}  總={stats['total_rows']:>5}  成功={stats['success']:>5}  空白={stats['skipped_empty']:>4}  資料過少={stats['skipped_thin']:>3}  錯={len(stats['errors'])}")
    finally:
        if db is not None:
            db.close()

    # 錯誤詳情
    print("\n" + "=" * 60)
    total_success = sum(s["success"] for s in all_stats)
    total_errors = sum(len(s["errors"]) for s in all_stats)
    print(f"✅ 總成功筆數: {total_success}")
    print(f"❌ 總錯誤筆數: {total_errors}")

    if total_errors:
        print("\n錯誤前 10 筆:")
        shown = 0
        for s in all_stats:
            for e in s["errors"]:
                if shown >= 10:
                    break
                print(f"  [{s['sheet']}] row {e['row']}: cert={e.get('cert_no')} — {e['error']}")
                shown += 1
            if shown >= 10:
                break

    if not args.commit:
        print("\n💡 確認以上數字 OK 後，加 --commit 真匯入。")


if __name__ == "__main__":
    main()
