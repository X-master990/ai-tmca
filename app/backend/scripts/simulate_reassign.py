"""不連 DB，直接在總表 xlsx 上預覽客戶編號「整批重配」結果（識別鍵 = 店名 + 地址）。

用途：在對線上 DB 跑 `assign_customer_no.py --reassign --apply` 之前，先離線看
新規則會把哪些「同店不同寫法」併成一號，以及哪些「同店名被地址拆成多號」需承辦複查。

直接套用 app/services/customer_no.py 的 holder_key / norm_address，與正式配號一致。

用法：python scripts/simulate_reassign.py <xlsx_path>
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from app.services.customer_no import (  # noqa: E402
    PREFIX_BY_CATEGORY,
    holder_key,
    norm_holder_name,
)
from scripts.import_excel import SHEET_CONFIG  # noqa: E402


def col_idx(cfg, field):
    for idx, _k, name, _d in cfg["columns"]:
        if name == field:
            return idx
    return None


def cell(row, idx):
    return str(row[idx - 1]).strip() if idx and idx - 1 < len(row) and row[idx - 1] is not None else ""


def main() -> None:
    if len(sys.argv) < 2:
        print("用法：python scripts/simulate_reassign.py <xlsx_path>")
        sys.exit(1)
    wb = load_workbook(sys.argv[1], read_only=True, data_only=True)

    groups: dict[tuple, str] = {}
    g_names: dict[tuple, set[str]] = defaultdict(set)
    g_addrs: dict[tuple, set[str]] = defaultdict(set)
    name_nos: dict[tuple, set[str]] = defaultdict(set)   # (prefix, norm_name) -> {customer_no}
    serial: dict[str, int] = defaultdict(int)
    old_keys: set[tuple] = set()
    assigned = no_name = no_prefix = with_addr = 0

    for sheet, cfg in SHEET_CONFIG.items():
        if sheet not in wb.sheetnames:
            continue
        prefix = PREFIX_BY_CATEGORY.get(cfg["category_code"])
        hc = col_idx(cfg, "holder_name")
        ac = col_idx(cfg, "use_address")
        tc = col_idx(cfg, "tax_id")
        for row in wb[sheet].iter_rows(min_row=cfg.get("data_start_row", 3), values_only=True):
            name = cell(row, hc)
            if not name:
                continue
            if not prefix:
                no_prefix += 1
                continue
            addr = cell(row, ac)
            hk = holder_key(name, addr)
            if hk is None:
                no_name += 1
                continue
            if addr:
                with_addr += 1
            key = (prefix, hk)
            if key not in groups:
                serial[prefix] += 1
                groups[key] = f"{prefix}{serial[prefix]:06d}"
            g_names[key].add(name)
            g_addrs[key].add(addr)
            name_nos[(prefix, norm_holder_name(name))].add(groups[key])
            old_keys.add((prefix, name, cell(row, tc)))
            assigned += 1

    wb.close()

    merges = [(k, g_names[k], g_addrs[k]) for k in groups
              if len(g_names[k]) > 1 or len(g_addrs[k]) > 1]
    name_split = sorted(((k, nos) for k, nos in name_nos.items() if len(nos) > 1),
                        key=lambda kv: -len(kv[1]))

    print(f"\n{'=' * 66}")
    print(f"客戶編號重配預覽（店名 + 地址）　指派 {assigned:,}　有地址 {with_addr:,}")
    print(f"{'-' * 66}")
    print(f"  舊規則 不重複號(店名+統編逐字): {len(old_keys):>5}")
    print(f"  新規則 不重複號(店名+地址)    : {len(groups):>5}")
    print("  各前綴店家數:")
    for p in sorted(serial):
        print(f"    {p}: {serial[p]} 家")
    print(f"{'-' * 66}")
    print(f"  併號群（同店不同寫法收成一號）: {len(merges)} 家")
    print(f"  同店名被地址拆成多號（請複查）: {len(name_split)} 個店名")
    print("  ※ 同名多址多為真不同分點（如音響廠商跨場地）；少數是地址打法不同的 under-merge，")
    print("    屬最佳努力，請承辦於系統內人工併號（寧分勿錯併）。")

    print("\n  —— 併號樣本（同一號收進多種店名/地址寫法，前 30）——")
    for (prefix, hk), names, addrs in sorted(merges, key=lambda kv: -(len(kv[1]) + len(kv[2])))[:30]:
        print(f"    {groups[(prefix, hk)]}  ←  {' ｜ '.join(sorted(n for n in names if n))}")

    print("\n  —— 同店名被拆成多號樣本（承辦複查：真不同分點? 還是地址打法不同?）前 30 ——")
    for (prefix, nn), nos in name_split[:30]:
        print(f"    [{prefix}] {nn}：{len(nos)} 號 {' '.join(sorted(nos))}")
    print(f"{'=' * 66}\n")


if __name__ == "__main__":
    main()
