"""回填：用修好的清洗函式，把當初被靜默轉成 NULL、現在能無歧義救回的值補進 DB。

設計原則（安全優先）:
  - 只「補空」：DB 該欄目前為 NULL、且重新清洗能得到非空值時才寫入。
  - 絕不覆蓋任何既有非空值（保護匯入後線上手動修改過的資料）。
  - 比對鍵 = (category_code, cert_no)，且 xlsx 與 DB 兩邊都唯一才動；否則跳過並列入人工清單。
  - 救不回的（一格兩值、不存在的日期、錯年份、非數字…）→ 全部寫進人工修正清單。

用法:
    # 乾跑（預設，不寫 DB），輸出將變更清單 + 人工清單
    python scripts/backfill_cleaned.py "/tmp/total.xlsx"
    # 實際寫入
    python scripts/backfill_cleaned.py "/tmp/total.xlsx" --apply -o IMPORT-MANUAL-FIXLIST.md
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import Record
from scripts.import_excel import (
    SHEET_CONFIG,
    _clean,
    _to_date,
    _to_int,
    _to_period_date,
    is_empty_row,
)

# 要回填的目標欄位中文標籤（除錯/清單顯示用）
LABELS = {
    "issued_date": "發證日", "invoice_date": "發票日期", "apply_date": "申請日期",
    "amount": "金額", "qty": "數量", "period_start": "授權起", "period_end": "授權迄",
    "audience_size": "人數", "floor_area": "坪數",
}


def _cell(row: tuple, idx1: int) -> Any:
    return row[idx1 - 1] if idx1 - 1 < len(row) else None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--apply", action="store_true", help="實際寫入 DB（預設乾跑）")
    ap.add_argument("-o", "--out", default="IMPORT-MANUAL-FIXLIST.md")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    db = SessionLocal()

    rescued: list[tuple] = []        # (sheet, row, cert_no, field, new_value)
    manual: list[tuple] = []         # (sheet, row, cert_no, label, raw, reason)
    skipped_match: list[tuple] = []  # (sheet, row, cert_no, reason)

    try:
        for sheet_name, cfg in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            code = cfg["category_code"]
            cols_cfg = cfg["columns"]
            start = cfg["data_start_row"]

            # xlsx 內 cert_no 出現次數（去重判斷）
            xlsx_cert_count: dict[str, int] = defaultdict(int)
            parsed_rows = []
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=start, values_only=True), start=start
            ):
                if is_empty_row(row):
                    continue
                cert = _clean(_cell(row, 4))  # col4 = 證書編號/證號
                parsed_rows.append((row_idx, row, cert))
                if cert:
                    xlsx_cert_count[str(cert)] += 1

            for row_idx, row, cert in parsed_rows:
                # 逐個目標欄位：偵測「原本有值卻清洗成 None」(人工) 與 可救的新值
                row_new: dict[str, Any] = {}
                row_manual: list[tuple] = []

                for col_idx, kind, key, dtype in cols_cfg:
                    if kind in ("ps_y", "ps_m", "ps_d", "pe_y", "pe_m", "pe_d"):
                        continue
                    if dtype not in ("date", "int"):
                        continue
                    raw = _cell(row, col_idx)
                    if _clean(raw) is None:
                        continue
                    new_val = _to_date(raw) if dtype == "date" else _to_int(raw)
                    if new_val is None:
                        row_manual.append((LABELS.get(key, key), repr(raw), "清洗後仍無法解析（歧義/不合法）"))
                    else:
                        row_new[(kind, key)] = new_val

                # 授權期間（y/m/d 三欄）
                ps = _to_period_date(_period(cols_cfg, row, "ps_y"), _period(cols_cfg, row, "ps_m"), _period(cols_cfg, row, "ps_d"))
                pe = _to_period_date(_period(cols_cfg, row, "pe_y"), _period(cols_cfg, row, "pe_m"), _period(cols_cfg, row, "pe_d"))
                ps_has = any(_clean(_period(cols_cfg, row, k)) is not None for k in ("ps_y", "ps_m", "ps_d"))
                pe_has = any(_clean(_period(cols_cfg, row, k)) is not None for k in ("pe_y", "pe_m", "pe_d"))
                if ps_has and ps is None:
                    row_manual.append(("授權起", _period_repr(cols_cfg, row, "ps"), "年/月/日組不成合法日期"))
                if pe_has and pe is None:
                    row_manual.append(("授權迄", _period_repr(cols_cfg, row, "pe"), "年/月/日組不成合法日期"))
                if ps is not None:
                    row_new[("col", "period_start")] = ps
                if pe is not None:
                    row_new[("col", "period_end")] = pe

                for label, raw, reason in row_manual:
                    manual.append((sheet_name, row_idx, cert or "", label, raw, reason))

                if not row_new:
                    continue

                # 比對鍵唯一才動
                if not cert:
                    skipped_match.append((sheet_name, row_idx, "", "無證書編號，無法定位"))
                    continue
                if xlsx_cert_count[str(cert)] != 1:
                    skipped_match.append((sheet_name, row_idx, cert, "xlsx 內證號重複"))
                    continue
                matches = (
                    db.query(Record)
                    .filter(Record.category_code == code, Record.cert_no == str(cert))
                    .all()
                )
                if len(matches) != 1:
                    skipped_match.append((sheet_name, row_idx, cert, f"DB 配對 {len(matches)} 筆"))
                    continue
                target = matches[0]

                for (kind, key), new_val in row_new.items():
                    if kind == "extra":
                        cur = (target.extra or {}).get(key)
                        if cur in (None, ""):
                            if args.apply:
                                ne = dict(target.extra or {})
                                ne[key] = new_val
                                target.extra = ne
                            rescued.append((sheet_name, row_idx, cert, f"extra.{key}", new_val))
                    else:
                        cur = getattr(target, key)
                        if cur is None:
                            if args.apply:
                                setattr(target, key, new_val)
                            rescued.append((sheet_name, row_idx, cert, key, new_val))

        if args.apply:
            db.commit()
    finally:
        db.close()

    # ---- 報表 ----
    print(f"\n{'=' * 60}")
    print(f"回填{'（已寫入 DB）' if args.apply else '（乾跑，未寫入）'}")
    print(f"  可救回（補空）: {len(rescued)} 個欄位值")
    by_field: dict[str, int] = defaultdict(int)
    for _, _, _, f, _ in rescued:
        by_field[f] += 1
    for f, n in sorted(by_field.items(), key=lambda x: -x[1]):
        print(f"    - {f}: {n}")
    print(f"  需人工修正: {len(manual)} 筆")
    print(f"  無法定位而跳過: {len(skipped_match)} 筆")
    print(f"{'=' * 60}\n")

    _write_manual(args.out, rescued, manual, skipped_match, applied=args.apply)
    print(f"📄 清單已寫入: {args.out}")


def _period(cols_cfg, row, kind):
    for col_idx, k, _, _ in cols_cfg:
        if k == kind:
            return _cell(row, col_idx)
    return None


def _period_repr(cols_cfg, row, prefix):
    ys = _period(cols_cfg, row, f"{prefix}_y")
    ms = _period(cols_cfg, row, f"{prefix}_m")
    ds = _period(cols_cfg, row, f"{prefix}_d")
    return f"年={ys!r} 月={ms!r} 日={ds!r}"


def _write_manual(path, rescued, manual, skipped, applied):
    lines = ["# 匯入失真 — 回填結果 & 人工修正清單", ""]
    lines.append(f"- 自動補回（{'已寫入' if applied else '乾跑'}）：**{len(rescued)}** 個欄位值")
    lines.append(f"- 需人工逐筆修正：**{len(manual)}** 筆（程式無法無歧義判定）")
    lines.append(f"- 無法定位跳過：**{len(skipped)}** 筆")
    lines.append("")
    lines.append("## 一、已自動補回的值（原為 NULL，現清洗成功）")
    lines.append("")
    lines.append("| sheet | 行 | 證號 | 欄位 | 補回值 |")
    lines.append("|---|---:|---|---|---|")
    for s, r, c, f, v in rescued:
        lines.append(f"| {s} | {r} | {c} | {f} | {v} |")
    lines.append("")
    lines.append("## 二、需人工修正（程式救不了，請逐筆判定後在系統上修改）")
    lines.append("")
    lines.append("| sheet | 行 | 證號 | 欄位 | 原始值 | 原因 |")
    lines.append("|---|---:|---|---|---|---|")
    for s, r, c, label, raw, reason in manual:
        lines.append(f"| {s} | {r} | {c} | {label} | `{raw}` | {reason} |")
    if skipped:
        lines.append("")
        lines.append("## 三、有可補值但無法定位 DB 紀錄（證號缺/重複）")
        lines.append("")
        lines.append("| sheet | 行 | 證號 | 原因 |")
        lines.append("|---|---:|---|---|")
        for s, r, c, reason in skipped:
            lines.append(f"| {s} | {r} | {c} | {reason} |")
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
