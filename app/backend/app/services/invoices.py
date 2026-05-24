"""發票生成 — 讀模板 → 填明細 → 配發票號

模板：/var/tmca/templates/發票/invoice_template.xlsx
  - Sheet 明細：A 發票型式, B 抬頭, C 統編, E 品名, F 未稅單價, I 數量,
                Q 開立日期, R 發票號碼, S 客戶編號, T 備註
  - Sheet 設置：B2=prefix, C2=起始號（顯示用，實際序號由 DB 管）
"""
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.models import InvoiceSequence, Record

TEMPLATE_PATH = Path(settings.template_dir) / "發票" / "invoice_template.xlsx"

# 12 個 category_code → 品名（依授權型態歸三類，其他歸權利金）
CATEGORY_TO_PRODUCT = {
    "PUBLIC_TRANSMIT":   "音樂公開傳輸授權費",
    "HALL_ROOM":         "音樂公開播送授權費",
    "AREA_DISPLAY":      "音樂公開播送授權費",
    "COMMUNITY_BOARD":   "音樂公開播送授權費",
    "TRANSPORT":         "音樂公開播送授權費",
    "COMPUTER_KARAOKE":  "音樂公開演出授權費",
    "PUBLIC_KARAOKE":    "音樂公開演出授權費",
    "SELF_SERVICE_KTV":  "音樂公開演出授權費",
    "STREET_ARTIST":     "音樂公開演出授權費",
    "SINGLE_EVENT":      "音樂公開演出授權費",
    "FUNERAL":           "音樂公開演出授權費",
    "ELECTION":          "音樂公開演出授權費",
}
DEFAULT_PRODUCT = "權利金收入"


def _allocate_numbers(db: Session, prefix: str, count: int) -> list[str]:
    """以 SELECT FOR UPDATE 鎖序號列、配 count 個號碼。"""
    seq = (
        db.query(InvoiceSequence)
        .filter(InvoiceSequence.prefix == prefix)
        .with_for_update()
        .one_or_none()
    )
    if seq is None:
        raise ValueError(f"invoice_sequence 找不到 prefix={prefix}")
    start = seq.next_no
    nums = [f"{prefix}{start + i:08d}" for i in range(count)]
    seq.next_no = start + count
    db.flush()
    return nums


def generate_invoice_xlsx(
    db: Session,
    record_ids: list[int],
    issue_date: date,
    invoice_type: str = "二聯式",
    prefix: str = "TU",
) -> tuple[bytes, list[tuple[int, str]]]:
    """填模板、配號，回 (xlsx_bytes, [(record_id, invoice_no), ...])"""
    if not record_ids:
        raise ValueError("record_ids 不能為空")
    if invoice_type not in ("二聯式", "三聯式", "作廢"):
        raise ValueError(f"invoice_type 不合法：{invoice_type}")

    recs = db.query(Record).filter(Record.id.in_(record_ids)).all()
    by_id = {r.id: r for r in recs}
    ordered = [by_id[i] for i in record_ids if i in by_id]
    if not ordered:
        raise ValueError("找不到任何指定的 record")

    # 避免重複開立
    already = [r.id for r in ordered if r.invoice_no]
    if already:
        raise ValueError(f"以下 record 已有發票號，不可重複開立：{already}")

    numbers = _allocate_numbers(db, prefix, len(ordered))

    wb = load_workbook(TEMPLATE_PATH, keep_vba=False)
    ws = wb["明細"]

    assigned: list[tuple[int, str]] = []
    for i, (rec, inv_no) in enumerate(zip(ordered, numbers)):
        row = 2 + i  # 第 1 列是表頭
        product = CATEGORY_TO_PRODUCT.get(rec.category_code, DEFAULT_PRODUCT)
        title = rec.invoice_title or rec.holder_name or rec.applicant_name or ""
        total = int(rec.amount or 0)
        unit_price = round(total / 1.05) if total else 0

        ws.cell(row=row, column=1,  value=invoice_type)          # A 發票型式
        ws.cell(row=row, column=2,  value=title)                 # B 抬頭
        ws.cell(row=row, column=3,  value=rec.tax_id or "")      # C 統編
        ws.cell(row=row, column=5,  value=product)               # E 品名
        ws.cell(row=row, column=6,  value=unit_price)            # F 未稅單價
        ws.cell(row=row, column=9,  value=1)                     # I 數量
        ws.cell(row=row, column=17, value=issue_date)            # Q 開立日期
        ws.cell(row=row, column=18, value=inv_no)                # R 發票號碼（蓋掉公式，寫死配到的號）
        ws.cell(row=row, column=19, value=rec.customer_no or str(rec.id))  # S 客戶編號（無號則退回 record id）
        ws.cell(row=row, column=20, value=rec.note or "")        # T 備註

        assigned.append((rec.id, inv_no))

    # 同步 設置!C2 顯示「下一張起始號」
    setup = wb["設置"]
    setup.cell(row=2, column=2, value=prefix)
    next_after = int(numbers[-1][len(prefix):]) + 1
    setup.cell(row=2, column=3, value=next_after)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), assigned
